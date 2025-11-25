import re
import argparse
from pathlib import Path
import openpyxl
import csv
from collections import defaultdict
import matplotlib.pyplot as plt
import numpy as np 

def set_nested_value(d, path, value):
    """Sets a value in a nested dictionary given a list of keys."""    
    
    key = path[0]
    if(len(path) == 1) : 
        if(key not in d) : 
            d[key] = []
        d[key].append(value)
    else : 
        if key not in d:
            d[key] = {}
        d[key] = set_nested_value(d[key], path[1:], value)
    return d

def extract_name_from_kernel_info(kernel_info_str):
    """Extract the 'name' field from the kernel info string."""
    try:
        # Parse the string to extract the name
        # The format is: "[{'name': '...', 'stream': ..., ...}]"
        if kernel_info_str is None or kernel_info_str == '':
            return None

        # Try to extract just the name using regex
        match = re.search(r"'name':\s*'([^']+)'", str(kernel_info_str))
        if match:
            return match.group(1)

        return None
    except Exception as e:
        print(f"Error parsing kernel info: {e}")
        return None

def column_letter_to_index(letter):
    """Convert Excel column letter to 0-based index."""
    index = 0
    for i, char in enumerate(reversed(letter.upper())):
        index += (ord(char) - ord('A') + 1) * (26 ** i)
    return index - 1


def process_excel_file(file_path, threads, channel, rank, top_k=5):
    """Process a single Excel file and extract GEMM data."""
    try:
        # Open the workbook
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        # Check if GEMM sheet exists
        if 'GEMM' not in wb.sheetnames:
            print(f"Warning: GEMM sheet not found in {file_path}")
            return []

        sheet = wb['GEMM']

        # Column indices (0-based)
        # X = 23 (24th column)
        # AG = 32 (33rd column) - Kernel Time - ms (min)
        # AH = 33 (34th column) - Kernel Time - max (max)
        col_x = column_letter_to_index('X')
        col_ag = column_letter_to_index('AG')
        col_ah = column_letter_to_index('AH')
        col_af = column_letter_to_index('AF')

        # Read all rows
        rows_data = []
        header_row = None

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                # This is the header
                header_row = list(row)
                continue

            if row is None or len(row) <= max(col_x, col_ag, col_ah):
                continue

            kernel_info = row[col_x] if col_x < len(row) else None
            kernel_time_min = row[col_ag] if col_ag < len(row) else None
            kernel_time_max = row[col_ah] if col_ah < len(row) else None
            kernel_time_std = row[col_af] if col_af < len(row) else None

            # Extract kernel name
            kernel_name = extract_name_from_kernel_info(kernel_info)

            # Calculate time difference
            if kernel_time_min is not None and kernel_time_max is not None:
                try:
                    time_diff = float(kernel_time_max) - float(kernel_time_min)
                except (ValueError, TypeError):
                    continue
            else:
                continue

            if kernel_name:
                row_dict = {
                    'threads': threads,
                    'channel': channel,
                    'rank': rank,
                    'kernel_name': kernel_name,
                    'kernel_time_min_ms': kernel_time_min,
                    'kernel_time_max_ms': kernel_time_max,
                    'kernel_time_std' : kernel_time_std,
                    'time_diff': time_diff
                }

                rows_data.append(row_dict)

        wb.close()

        # Sort by time_diff and get top k
        #rows_data.sort(key=lambda x: x['time_diff'], reverse=True)
        #top_results = rows_data[:top_k]

        return rows_data

    except Exception as e:
        print(f"Error processing {file_path}: {e}")
        import traceback
        traceback.print_exc()
        return []

def parse_args():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="Analyze GEMM reports and extract top kernels with highest time variance",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Use default settings
  python analyze_gemm_reports.py
  # Specify custom base path
  python analyze_gemm_reports.py --base-path /path/to/tracelens_analysis
  # Specify custom configurations
  python analyze_gemm_reports.py --threads 256 512 --channels 28 42 56 70 --ranks 0 1 2 3 4 5 6 7
  # Extract top 10 kernels instead of top 5
  python analyze_gemm_reports.py --top-k 10
  # Custom output file
  python analyze_gemm_reports.py --output-plot-directory path/to/directory/to/save/plots
        """
    )

    parser.add_argument(
        '--base-path',
        type=Path,
        default=Path("/home/oyazdanb/aorta/experiments/sweep_20251121_155219/tracelens_analysis"),
        help='Base path to tracelens_analysis directory (default: %(default)s)'
    )

    parser.add_argument(
        '--threads',
        type=int,
        nargs='+',
        default=[256, 512],
        help='Thread configurations to analyze (default: %(default)s)'
    )

    parser.add_argument(
        '--channels',
        type=int,
        nargs='+',
        default=[28, 42, 56, 70],
        help='Channel configurations to analyze (default: %(default)s)'
    )

    parser.add_argument(
        '--ranks',
        type=int,
        nargs='+',
        default=list(range(8)),
        help='Ranks to analyze (default: 0 1 2 3 4 5 6 7)'
    )

    parser.add_argument(
        '--top-k',
        type=int,
        default=5,
        help='Number of top kernels to extract per file (default: %(default)s)'
    )

    parser.add_argument(
        '-o',
        '--output-plot-directory',
        type=Path,
        default='/home/oyazdanb/aorta/experiments/sweep_20251121_155219/tracelens_analysis/plots',
        help='Output CSV filename (default: %(default)s)'
    )


    return parser.parse_args()

def extract_data_from_top_kernels(all_results, top_kernels) : 
    
    kernel_plot_data = defaultdict(lambda: defaultdict(list))
    
    
    for kernel_name in top_kernels :     
        kernel_data = [r for r in all_results if r['kernel_name'] == kernel_name ]
        for k in kernel_data : 
            kernel_plot_data[k['kernel_name']][f"t{k['threads']}_ch{k['channel']}_r{k['rank']}"].append(k['time_diff'])
    
    return kernel_plot_data

def plot_top_kernels(all_results, top_k, output_dir, subplots=False) :

    all_results.sort(key=lambda x: x['time_diff'], reverse=True)
    top_kernels = [row['kernel_name'] for row in all_results[:top_k] ]
    unique_top_kernel = list(dict.fromkeys(top_kernels))

    if(len(top_kernels) != len(unique_top_kernel)) :
        print("Duplicate top kernel, please reurn using higher top-k value")

    print("Top kernel list : ")
    for kernel in unique_top_kernel : 
        print(f"\t {unique_top_kernel.index(kernel)}. {kernel}")

    #generate_kernel_analysis(all_results, unique_top_kernel)
    
    kernel_plot_data  = extract_data_from_top_kernels(all_results, unique_top_kernel)

    for kernel_name, kernel_info in kernel_plot_data.items() :
        x_data_label = [key for key, value in kernel_info.items()]
        
        x_data = range(len(x_data_label))
        y_data = []
        for key, value in kernel_info.items() :
            if(subplots == True) :
                if(len(value) > 1) : 
                    run_x_data = range(len(value))
                    run_y_data = value
                    #print(f"{kernel_name} {key} {run_x_data} {run_y_data}")
                    plt.figure()
                    plt.bar(run_x_data, run_y_data)
                    plt.xlabel("Occurance")
                    plt.ylabel("Time(ms)")
                    plt.tight_layout()
                    plt.savefig(f"Kernel_{unique_top_kernel.index(kernel_name)}{key}")
                    plt.close()
            y_data.append(max(value))
        
        plt.figure()
        plt.bar(x_data, y_data, width=0.25) 
        plt.title(f'Kernel {unique_top_kernel.index(kernel_name)}')
        plt.xticks(x_data, x_data_label, rotation=90, ha='right', fontsize=6)
        plt.ylabel("Tims(ms)")
        plt.tight_layout()
        plt.savefig(output_dir / f"Kernel_{unique_top_kernel.index(kernel_name)}")
        plt.close()
 
def main():
    # Parse command line arguments
    args = parse_args()

    base_path = args.base_path
    thread_configs = args.threads
    channels = args.channels
    ranks = args.ranks
    top_k = args.top_k

    output_dir = args.output_plot_directory 
    output_dir.mkdir(exist_ok=True, parents=True)

    # Validate base path
    if not base_path.exists():
        print(f"Error: Base path does not exist: {base_path}")
        return

    print(f"Configuration:")
    print(f"  Base path: {base_path}")
    print(f"  Threads: {thread_configs}")
    print(f"  Channels: {channels}")
    print(f"  Ranks: {ranks}")
    print(f"  Top K: {top_k}")
    print(f"  Output plot directory: {output_dir}")
    print()

    all_results = []

    print("Processing Excel files...")
    total_files = len(thread_configs) * len(channels) * len(ranks)
    file_count = 0

    for threads in thread_configs:
        thread_dir = base_path / f"{threads}thread" / "individual_reports"

        for channel in channels:
            for rank in ranks:
                file_name = f"perf_{channel}ch_rank{rank}.xlsx"
                file_path = thread_dir / file_name

                file_count += 1
                print(f"Processing {file_count}/{total_files}: {file_name}")

                if not file_path.exists():
                    print(f"  Warning: File not found: {file_path}")
                    continue

                # Process the file
                results = process_excel_file(file_path, threads, channel, rank, top_k)

                if results:
                    all_results.extend(results)
                    print(f"  Found {len(results)} kernels")
    
    plot_top_kernels(all_results=all_results, output_dir=output_dir, top_k=top_k)
    

    

if __name__ == "__main__":
    main()