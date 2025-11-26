import argparse
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook
from statistics import geometric_mean
import matplotlib.pyplot as plt
import numpy as np 

def process_comparison_data(base_path, channel, thread, rank) : 
    all_results = defaultdict(lambda :  defaultdict(list))
    #all_results['busy_time']["256"]["0"].append(10)
    #all_results['busy_time']["256"]["0"].append(29)

    comparison_path = base_path / "comparisons"

    col_type_name = 0 
    col_256_time = 3
    col_512_time = 5 

    for t in thread : 
        for c in channel : 
            for r in rank : 
                file_name = f"compare_{c}ch_rank{r}_across_threads.xlsx"
                file_path = comparison_path / file_name
                
                if not file_path.exists():
                    print(f"  Warning: File not found: {file_path}")
                    continue
                
                workbook = load_workbook(file_path)
                gpu_sheet = workbook['gpu_timeline']

                for i, row in enumerate(gpu_sheet.iter_rows(values_only=True)):
                    if i==0 : 
                        continue 
                    all_results[row[col_type_name]]['256'].append(float(row[col_256_time]))
                    all_results[row[col_type_name]]['512'].append(float(row[col_512_time]))
    print("Done reading excels.")
    return all_results 
def get_thread_and_type_values_over_ranks_with_mean_channel(all_result) :
    #mean_result =  defaultdict(lambda : defaultdict(list))
    type_list = [] 
    mean_result = defaultdict(list) 
    for type, type_info in all_result.items() :
        type_list.append(type)
        for t_id, time_info in type_info.items() :
            mean_result[t_id].append(geometric_mean(time_info))
    print("Done computing geomeans across channels.")
    return type_list, mean_result

def plot_mean_result(output_dir, types, mean_result, threads) :
    bar_width = 0.35
    x_pos = np.arange(len(types))
    output_file = output_dir / "comparison_summary.png"
    
    plt.figure()
    
    if(len(threads) > 1)  : 
        plt.bar(x_pos - (bar_width/2), mean_result[str(threads[0])], bar_width, label="256", color='r')
        plt.bar(x_pos + (bar_width/2), mean_result[str(threads[1])], bar_width, label="512", color='b')
    else :
        plt.bar(x_pos, mean_result[str(threads[0])], bar_width, label="256", color='b')
    plt.ylabel("Time")
    plt.xlabel("GPU Component")
    plt.xticks(x_pos, types, rotation=45, ha='right', fontsize=14)
    plt.title('GPU Component Summary v/s Time')
    plt.legend()
    plt.tight_layout()
    plt.savefig(output_file)
    plt.close()
    print("Done plotting.")

def parse_args():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="Analyze tracelens comparison report for components",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Use default settings
  python analyze_gemm_reports.py
  # Specify custom base path
  python analyze_gemm_reports.py --base-path /path/to/tracelens_analysis
  # Specify custom configurations
  python analyze_gemm_reports.py --threads 256 512 --channels 28 42 56 70 --ranks 0 1 2 3 4 5 6 7
  # Extract top 10 kernels instead of top 5
  python analyze_gemm_reports.py --top-k 10
  # Custom output file
  python analyze_gemm_reports.py --output-plot-directory path/to/directory/to/save/plots
        """
    )

    parser.add_argument(
        '--base-path',
        type=Path,
        default=Path("/home/oyazdanb/aorta/experiments/sweep_20251121_155219/tracelens_analysis"),
        help='Base path to tracelens_analysis directory (default: %(default)s)'
    )

    parser.add_argument(
        '--threads',
        type=int,
        nargs='+',
        default=[256, 512],
        help='Thread configurations to analyze (default: %(default)s)'
    )

    parser.add_argument(
        '--channels',
        type=int,
        nargs='+',
        default=[28, 42, 56, 70],
        help='Channel configurations to analyze (default: %(default)s)'
    )

    parser.add_argument(
        '--ranks',
        type=int,
        nargs='+',
        default=list(range(8)),
        help='Ranks to analyze (default: 0 1 2 3 4 5 6 7)'
    )

    parser.add_argument(
        '-o',
        '--output-plot-directory',
        type=Path,
        default=None,
        help='Output CSV filename (default: %(default)s)'
    )

    parser.add_argument(
        '--types',
        type=str,
        nargs='+',
        default=['busy_time', 'computation_time', 'exposed_comm_time']
    )


    return parser.parse_args()
def main():
    # Parse command line arguments
    args = parse_args()

    base_path = args.base_path
    thread_configs = args.threads
    channels = args.channels
    ranks = args.ranks
    

    output_dir = args.output_plot_directory
    if(args.output_plot_directory is None) : 
       output_dir =  base_path / "plots"
     
    output_dir.mkdir(exist_ok=True, parents=True)

    # Validate base path
    if not base_path.exists():
        print(f"Error: Base path does not exist: {base_path}")
        return

    print(f"Configuration:")
    print(f"  Base path: {base_path}")
    print(f"  Threads: {thread_configs}")
    print(f"  Channels: {channels}")
    print(f"  Ranks: {ranks}")
    print(f"  Output plot directory: {output_dir}")

    all_results = process_comparison_data(base_path=base_path, channel=channels, thread=thread_configs, rank=ranks)
    type_list, mean_results = get_thread_and_type_values_over_ranks_with_mean_channel(all_results)
    plot_mean_result(output_dir, type_list, mean_results, threads=thread_configs)

if __name__ == "__main__":
    main()

    
